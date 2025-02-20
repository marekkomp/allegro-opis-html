import streamlit as st
import pandas as pd
from bs4 import BeautifulSoup
import io
import json
from openpyxl import load_workbook

# Funkcja czyszcząca HTML i usuwająca niepotrzebne fragmenty JSON
def clean_html(content):
    # Usuwamy fragmenty JSON-like, zachowując tylko dane zawarte w "content"
    try:
        data = json.loads(content)  # Próbujemy zinterpretować zawartość jako JSON
        # Sprawdzamy, czy to jest lista "items" i wyciągamy tylko zawartość typu "TEXT"
        clean_content = ""
        for section in data.get("sections", []):
            for item in section.get("items", []):
                if item.get("type") == "TEXT" and "content" in item:
                    clean_content += item["content"]
        return clean_content
    except json.JSONDecodeError:
        # Jeśli nie jest to poprawny JSON, po prostu zwracamy oryginalny tekst
        return content

# Funkcja do wczytania pliku Excel z pominięciem ukrytych wierszy
def read_excel_skip_hidden(uploaded_file):
    # Wczytanie pliku za pomocą openpyxl
    wb = load_workbook(uploaded_file, data_only=True)
    sheet = wb.active  # Wybieramy aktywny arkusz

    # Pobieramy wszystkie widoczne wiersze
    rows = []
    for row in sheet.iter_rows(min_row=4):  # Rozpoczynamy od 4. wiersza, aby pominąć nagłówek
        if not sheet.row_dimensions[row[0].row].hidden:  # Sprawdzamy, czy wiersz jest widoczny
            rows.append([cell.value for cell in row])

    # Tworzymy DataFrame z danych, z pominięciem ukrytych wierszy
    df = pd.DataFrame(rows, columns=[cell.value for cell in sheet[4]])  # Używamy 4. wiersza jako nagłówka
    return df

# Streamlit UI
st.title("Przetwarzanie plików Excel i czyszczenie HTML w opisach ofert")

# Wgrywanie pliku Excel
uploaded_file = st.file_uploader("Wybierz plik Excel", type=["xlsm", "xlsx"])

if uploaded_file is not None:
    # Wczytanie pliku Excel z pominięciem ukrytych wierszy
    df = read_excel_skip_hidden(uploaded_file)

    # Sprawdzenie, czy kolumna "Opis oferty" istnieje w DataFrame
    if "Opis oferty" in df.columns:
        st.write("Oryginalne dane:")
        st.write(df[["Opis oferty"]].head())

        # Przetwarzanie kolumny "Opis oferty"
        df['Opis oferty'] = df['Opis oferty'].apply(clean_html)

        # Pokazanie przetworzonych danych
        st.write("Przetworzone dane:")
        st.write(df[["Opis oferty"]].head())

        # Przygotowanie pliku do pobrania
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Przetworzone oferty")
        output.seek(0)

        # Opcja pobrania poprawionego pliku
        st.download_button(
            label="Pobierz poprawiony plik Excel",
            data=output,
            file_name="poprawiony_oferty.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.error("Brak kolumny 'Opis oferty' w pliku Excel.")
